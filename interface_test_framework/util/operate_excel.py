# 安装openpyxl库： pip install openpyxl
from openpyxl import load_workbook

class OperateExcel:

    # 将操作excel的位置赋值给一个实例属性
    def __init__(self, fn = None):
        # 如果用户没有传入测试数据文件的位置，则给一个默认值
        if not fn:
            self.fn = '../test_data/testcases.xlsx'
        else:
            # 否则就使用用户传入的值
            self.fn = fn
        self.wb = load_workbook(self.fn)
        self.ws = self.get_data()

    # 打开excel文件并返回当前active状态的工作薄
    def get_data(self):
        ws = self.wb.active
        return ws

    # 获取所有行的行数
    def get_lines(self):
        return len(list(self.ws.rows))

    # 获取每个单元格的内容, 行和列的索引均为1开始
    def get_value(self, row, col):
        return self.ws.cell(row, col).value

    # 将结果写入excel中
    def write_result(self, row, col, value):
        self.ws.cell(row, col).value = value
        self.wb.save(self.fn)



if __name__ == '__main__':
    op = OperateExcel()
    print(op.get_value(4, 3))


